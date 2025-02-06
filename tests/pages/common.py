import pathlib

import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


@pytest.fixture(scope="module")
def setup_driver():
    # chrome_options = Options()
    # chrome_options.add_argument("--no-sandbox")
    # chrome_options.add_argument("--disable-dev-shm-usage")
    # chrome_options.add_argument("--headless")
    # chrome_options.add_argument("--disable-gpu")
    # driver = webdriver.Chrome(options=chrome_options)
    driver = webdriver.Chrome()
    driver.get('https://healthapp.yaksha.com/')
    driver.implicitly_wait(60)
    driver.maximize_window()
    yield driver
    driver.quit()


@pytest.fixture(scope="session")
def credentials():
    """
       Reads test credentials (username and password) from an Excel file.

       Returns:
           dict: A dictionary containing the test credentials.
       """
    file = pathlib.Path("./testData/Verification.xlsx")
    wb = openpyxl.load_workbook(file)
    sheet = wb["Credentials"]
    data = {}

    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = row[0]
        value = row[1]
        data[key] = value
    return data


@pytest.fixture(scope="session")
def expected_data():
    """
        Reads expected test data for verification operations from an Excel file.

        Returns:
            dict: A dictionary containing the expected test data.
    """
    file = pathlib.Path("./testData/Verification.xlsx")
    wb = openpyxl.load_workbook(file)
    sheet = wb["Verification"]
    data = {}

    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = row[0]
        value = row[1]
        data[key] = value

    return data